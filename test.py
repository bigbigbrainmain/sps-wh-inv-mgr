from tkinter import *
import tkinter as tk
from tkcalendar import DateEntry
from tkinter import ttk
from tkcalendar import Calendar
from tkinter import messagebox
from openpyxl import load_workbook
import datetime
import traceback
from datetime import date, datetime, timedelta
import re
from collections import deque

student_var = None
instrument_var = None
date_of_loan_entry = None
duration_of_loan_entry = None
duration_unit_var = None
cost_entry = None
confirmation_label = None  
page_stack = deque()

# Load the workbook
wb = load_workbook('database.xlsx')
student_sheet = wb['Student']
loans_sheet = wb['Current_Loans']
instruments_sheet = wb['Instruments']


def resize_buttons(root, buttons):
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    padding = 20
    button_width = (window_width - padding * 6) // 2  
    button_height = (window_height - padding * 3) // 3  

    buttons[0].place(x=padding, y=padding, width=button_width, height=button_height)
    buttons[1].place(x=window_width // 2 + padding // 2, y=padding, width=button_width, height=button_height)
    buttons[2].place(x=padding, y=window_height // 3 + padding // 2, width=button_width, height=button_height)
    buttons[3].place(x=window_width // 2 + padding // 2, y=window_height // 3 + padding // 2, width=button_width, height=button_height)  
    buttons[4].place(x=padding, y=window_height * 2 // 3 + padding // 2, width=2.19*button_width, height=button_height)  

def homepage(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(homepage)

    root.title("Wathen Inventory Manager")
    root.configure(background="#E2DEDD")
    root.minsize(640, 600)
    root.maxsize(1920, 1080)

    inventory_in = Button(root, text="IN", bg="green", fg="white", font=("Arial", 24), command=lambda: show_in_interface(root))
    inventory_out = Button(root, text="OUT", bg="red", fg="white", font=("Arial", 24), command=lambda: show_out_interface(root))
    inventory_info = Button(root, text="INFORMATION", bg="blue", fg="white", font=("Arial", 24), command=lambda: show_info_interface(root))
    inventory_modify = Button(root, text="MODIFY", bg="#F0EBBB", fg="black", font=("Arial", 24), command=lambda: show_modify_interface(root))
    inventory_returns = Button(root, text="RETURNS", bg="purple", fg="white", font=("Arial", 24), command=lambda: show_returns_interface(root))  


    buttons = [inventory_in, inventory_out, inventory_info, inventory_modify, inventory_returns]
    root.bind("<Configure>", lambda event: resize_buttons(root, buttons))
    resize_buttons(root, buttons)  # Initial button placement
        
def show_in_interface(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_in_interface)

    back_button = create_back_button(root, lambda: go_back(root))

    def submit_data_in():
        try:
            instrument_name = instrument_name_entry.get().strip()
            serial_no = serial_no_entry.get().strip()
            description = description_text.get("1.0", END).strip()
            date_of_purchase = date_of_purchase_entry.get()
            vendor_name = vendor_name_entry.get().strip()
            notes = notes_text.get("1.0", END).strip()
            maintenance_notes = maintenance_notes_text.get("1.0", END).strip()
            storage_location = storage_location_entry.get().strip()
            school = school_var.get()
            category = category_listbox.get(category_listbox.curselection())

            if not (instrument_name and description and date_of_purchase and vendor_name and storage_location and school and category and serial_no):
                mandatory_fields_info.config(text="Please fill in all mandatory fields.", fg="red")
                return

            # Load the workbook
            wb = load_workbook('database.xlsx')
            sheet = wb['Instruments']

            # Find the maximum ID value in the existing records
            max_id = 0
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[1] is not None and isinstance(row[1], (int, float)):
                    max_id = max(max_id, int(row[1]))

            # Set the next available ID
            next_id = max_id + 1

            # Append the data to the sheet
            sheet.append([
                instrument_name,                        #instrument name
                next_id,                                #instrument ID
                description,                            #instrument desc
                serial_no,                              #instrument serial
                "",                                     #date of last valuation
                "",                                     #last valuation amount
                "",                                     #last repair date
                date_of_purchase,                       #purchased date
                vendor_name,                            #vendor
                notes,                                  #notes
                maintenance_notes,                      #maintenance notes
                "",                                     #currently hired by
                1,                                      #is available
                storage_location,                       #storage location
                school,                                 #school
                category                                #category
            ])

            # Save the workbook
            wb.save('database.xlsx')

            # Update the "* Fields are mandatory" label with a success message
            mandatory_fields_info.config(text="Item added successfully! ✓", fg="green")
            root.after(5000, reset_mandatory_label)

            # Clear the input fields
            instrument_name_entry.delete(0, END)
            serial_no_entry.delete(0, END)
            description_text.delete("1.0", END)
            date_of_purchase_entry.delete(0, END)
            vendor_name_entry.delete(0, END)
            notes_text.delete("1.0", END)
            maintenance_notes_text.delete("1.0", END)
            storage_location_entry.delete(0, END)
            school_var.set("JUNIORS")
            category_listbox.selection_clear(0, END)
            category_listbox.selection_set(0)

            # Update the ID display
            id_display.config(text=str(next_id + 1))

        except PermissionError:
            messagebox.showerror("Error", "Cannot add new items. Please close the database file (database.xlsx) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def reset_mandatory_label():
        if mandatory_fields_info.winfo_exists():
            mandatory_fields_info.config(text="* Fields are mandatory", fg="black")

    def cycle_school():
        schools = ["JUNIORS", "SENIORS"]
        current_index = schools.index(school_var.get())
        next_index = (current_index + 1) % len(schools)
        school_var.set(schools[next_index])

    # Labels and Entry Widgets
    input_frame = Frame(root, bg="#E2DEDD")
    input_frame.pack(pady=20, padx=20)

    # Column 1
    col1 = Frame(input_frame, bg="#E2DEDD")
    col1.pack(side=LEFT, padx=20)

    instrument_name_label = Label(col1, text="Instrument Name: *", bg="#E2DEDD", font=("Arial", 18))
    instrument_name_label.pack(pady=(0, 5))
    instrument_name_entry = Entry(col1, font=("Arial", 18), justify='center')
    instrument_name_entry.pack(pady=5)

    id_frame = Frame(col1, bg="#E2DEDD")
    id_frame.pack(pady=(20, 5))

    id_label = Label(id_frame, text="ID:", bg="#E2DEDD", font=("Arial", 18))
    id_label.pack(side=LEFT, padx=5)

    # Load the workbook
    wb = load_workbook('database.xlsx')
    sheet = wb['Instruments']

    # Find the maximum ID value in the existing records
    max_id = 0
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1] is not None and isinstance(row[1], (int, float)):
            max_id = max(max_id, int(row[1]))

    # Set the next available ID
    next_id = max_id + 1
    id_display = Label(id_frame, text=str(next_id), bg="#E2DEDD", font=("Arial", 18))
    id_display.pack(side=LEFT, padx=5)

    description_label = Label(col1, text="Description: *", bg="#E2DEDD", font=("Arial", 18))
    description_label.pack(pady=(20, 5))
    description_text = Text(col1, font=("Arial", 18), height=3, width=30)
    description_text.pack(pady=5)

    date_of_purchase_label = Label(col1, text="Date of Purchase: *", bg="#E2DEDD", font=("Arial", 18))
    date_of_purchase_label.pack(pady=(20, 5))
    date_of_purchase_entry = DateEntry(col1, font=("Arial", 18), justify='center', date_pattern='dd/MM/yyyy')
    date_of_purchase_entry.pack(pady=5)

    school_label = Label(col1, text="School: *", bg="#E2DEDD", font=("Arial", 18))
    school_label.pack(pady=(20, 5))

    school_var = StringVar(value="JUNIORS")
    school_button = Button(col1, textvariable=school_var, command=cycle_school, font=("Arial", 18), justify='center', width=20)
    school_button.pack(pady=5)

    category_label = Label(col1, text="Category: *", bg="#E2DEDD", font=("Arial", 18))
    category_label.pack(pady=(20, 5))

    categories = ["String", "Woodwind", "Brass", "Percussion", "Keyboard", "Other"]
    category_listbox = Listbox(col1, font=("Arial", 18), height=3, exportselection=False)
    category_listbox.pack(pady=5)
    for category in categories:
        category_listbox.insert(END, category)
    category_listbox.selection_set(0)

    # Column 2
    col2 = Frame(input_frame, bg="#E2DEDD")
    col2.pack(side=LEFT, padx=20)

    serial_no_label = Label(col2, text="Serial No: *", bg="#E2DEDD", font=("Arial", 18))
    serial_no_label.pack(pady=(0, 5))
    serial_no_entry = Entry(col2, font=("Arial", 18), justify='center')
    serial_no_entry.pack(pady=5)

    vendor_name_label = Label(col2, text="Vendor Name: *", bg="#E2DEDD", font=("Arial", 18))
    vendor_name_label.pack(pady=(20, 5))
    vendor_name_entry = Entry(col2, font=("Arial", 18), justify='center')
    vendor_name_entry.pack(pady=5)

    notes_label = Label(col2, text="Notes:", bg="#E2DEDD", font=("Arial", 18))
    notes_label.pack(pady=(20, 5))
    notes_text = Text(col2, font=("Arial", 18), height=3, width=30)
    notes_text.pack(pady=5)

    maintenance_notes_label = Label(col2, text="Maintenance Notes:", bg="#E2DEDD", font=("Arial", 18))
    maintenance_notes_label.pack(pady=(20, 5))
    maintenance_notes_text = Text(col2, font=("Arial", 18), height=3, width=30)
    maintenance_notes_text.pack(pady=5)

    storage_location_label = Label(col2, text="Storage Location: *", bg="#E2DEDD", font=("Arial", 18))
    storage_location_label.pack(pady=(20, 5))
    storage_location_entry = Entry(col2, font=("Arial", 18), justify='center')
    storage_location_entry.pack(pady=5)    

    try:
        # Code that might raise an exception
        pass
    except PermissionError:
        messagebox.showerror("Error", "Cannot add new items. Please close the database file (database.xlsx) and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    # Submit Button
    submit_button = Button(root, text="Submit", font=("Arial", 18), command=submit_data_in)
    submit_button.pack(pady=20)

    # Mandatory fields information
    mandatory_fields_info = Label(root, text="* Fields are mandatory", bg="#E2DEDD", font=("Arial", 14))
    mandatory_fields_info.pack(pady=(0, 20))

    def reposition_submit_button(event):
        window_width = root.winfo_width()
        window_height = root.winfo_height()
        button_width = submit_button.winfo_width()
        button_height = submit_button.winfo_height()
        x = (window_width - button_width) // 2
        y = window_height - button_height - 40  # Adjust the value to position the button above the bottom
        submit_button.place(x=x, y=y)

    
    def raise_back_button(event):
        back_button.lift()

    root.bind("<Configure>", raise_back_button)


    def validate_required_fields(*args):
        if (
            instrument_name_entry.get()
            and description_text.get("1.0", END).strip()
            and date_of_purchase_entry.get()
            and vendor_name_entry.get()
            and storage_location_entry.get()
            and serial_no_entry.get()
            and school_var.get()
            and category_listbox.curselection()
        ):
            submit_button.config(state="normal")
        else:
            submit_button.config(state="disabled")

    # Call validation function initially
    validate_required_fields()

    # Bind validation function to Entry widgets
    instrument_name_entry.bind("<KeyRelease>", validate_required_fields)
    description_text.bind("<KeyRelease>", validate_required_fields)
    date_of_purchase_entry.bind("<KeyRelease>", validate_required_fields)
    vendor_name_entry.bind("<KeyRelease>", validate_required_fields)
    storage_location_entry.bind("<KeyRelease>", validate_required_fields)
    category_listbox.bind("<<ListboxSelect>>", validate_required_fields)

def clear_confirmation_label():
    if confirmation_label.winfo_exists():
        confirmation_label.config(text="")

def show_out_interface(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_out_interface)
    back_button = create_back_button(root, lambda: homepage(root))
    
    global student_var, instrument_var, date_of_loan_entry, duration_of_loan_entry, duration_unit_var, cost_entry, confirmation_label, selected_student_id
    selected_student_id = ""
    school_var = StringVar(value="JUNIORS")

    def read_term_dates():
        term_dates = []
        with open("term_dates.txt", "r") as file:
            for line in file:
                print(f"Line: {line}")
                line = line.strip()
                if line:  # Skip empty lines
                    year, term_name, start_date, end_date = line.split(",")
                    term_dates.append((int(year.strip()), term_name.strip(), datetime.strptime(start_date.strip(), "%d/%m/%Y"), datetime.strptime(end_date.strip(), "%d/%m/%Y")))
        return term_dates

    def get_current_term(current_date):
        current_datetime = datetime.combine(current_date, datetime.min.time())
        print(f"Current Date: {current_date}")
        print(f"Current Datetime: {current_datetime}")
        for term in term_dates:
            print(f"Term: {term}")
            if term[2] <= current_datetime <= term[3]:
                print(f"Matching Term: {term}")
                return term
        print("No matching term found")
        return None
    
    term_dates = read_term_dates()
    
    def toggle_school():
        schools = ["JUNIORS", "SENIORS"]
        current_index = schools.index(school_var.get())
        next_index = (current_index + 1) % len(schools)
        school_var.set(schools[next_index])
        update_instrument_search_list()

    def submit_data_out():
        student_name = student_var.get()
        date_of_loan = date_of_loan_entry.get_date()
        duration_value = duration_of_loan_entry.get()
        duration_unit = duration_unit_var.get()
        cost = cost_entry.get()
        instrument = instrument_var.get()

        if not student_name or not date_of_loan or not duration_value or not cost or not instrument:
            confirmation_label.config(text="Please fill in all the required fields.", fg="red")
            return

        wb = load_workbook('database.xlsx')
        loans_sheet = wb['Current_Loans']
        instruments_sheet = wb['Instruments']
        student_sheet = wb['Student']

        print("Workbook and sheets loaded successfully")

        max_row = loans_sheet.max_row
        loan_id = f"{max_row + 1}"
        instrument_id = instrument.split(" - ID: ")[1] if " - ID: " in instrument else ""
        date_of_loan_formatted = date_of_loan.strftime("%d%m%Y")

        current_date = datetime.now().date()
        current_term = get_current_term(current_date)

        if int(duration_value) == 0:
            if current_term:
                return_date = current_term[3].strftime("%d%m%Y")
            else:
                messagebox.showerror("Error", "No matching term found for current date.")
                return
        else:
            if duration_unit == "Half Terms":
                next_term_index = term_dates.index(current_term) + int(duration_value)
                if next_term_index < len(term_dates):
                    return_date = term_dates[next_term_index][3].strftime("%d%m%Y")
                else:
                    messagebox.showerror("Error", "Loan duration exceeds available term data.")
                    return
            elif duration_unit == "Terms":
                next_term_index = term_dates.index(current_term) + (int(duration_value) * 2)
                if next_term_index < len(term_dates):
                    return_date = term_dates[next_term_index][3].strftime("%d%m%Y")
                else:
                    messagebox.showerror("Error", "Loan duration exceeds available term data.")
                    return
            elif duration_unit == "Years":
                next_term_index = term_dates.index(current_term) + (int(duration_value) * 6)
                if next_term_index < len(term_dates):
                    return_date = term_dates[next_term_index][3].strftime("%d%m%Y")
                else:
                    last_term_date = term_dates[-1][3]
                    calculated_return_date = last_term_date + timedelta(days=365*int(duration_value))
                    if calculated_return_date <= term_dates[-1][3]:
                        return_date = calculated_return_date.strftime("%d%m%Y")
                    else:
                        messagebox.showerror("Error", "Loan duration exceeds available term data.")
                        return
            else:
                messagebox.showerror("Error", "Unsupported duration unit.")
                
                show_out_interface()
        new_row = [
            selected_student_id,
            instrument_id,
            date_of_loan_formatted,
            f"{duration_value} {duration_unit}",
            cost,
            return_date,
            ""
        ]
        print(f"New Row: {new_row}")
        loans_sheet.append(new_row)

        duration_value = duration_of_loan_entry.get()
        print(f"Duration Value: {duration_value}")

        # Update the "Instruments" sheet to mark the instrument as unavailable
        for row in range(2, instruments_sheet.max_row + 1):
            if str(instruments_sheet.cell(row=row, column=2).value) == instrument_id:
                instruments_sheet.cell(row=row, column=13, value=0)  # Set "Available" to 0
                instruments_sheet.cell(row=row, column=12, value=f"{student_name} - ID: {selected_student_id}")  # Set "Current Owner" to student info
                break

        student_info = student_name.split(" - ID: ")[0]
        preferred_name, surname, form = student_info.split(" ", 2)
        form = form.strip("()")
        print(f"Student info extracted: {preferred_name} {surname} - Form: {form}")

        student_found = False
        sheets = ['Divisions_spj', 'Divisions_sps']
        for sheet_name in sheets:
            print(f"Searching in sheet: {sheet_name}")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                forename, preferred, sur, student_form, _, school_id, year = row[:7]
                if preferred == preferred_name and sur == surname and student_form == form:
                    print(f"Student found in {sheet_name}: {preferred} {surname}")
                    school = "SPJ" if sheet_name == 'Divisions_spj' else "SPS"
                    new_row = student_sheet.max_row + 1
                    student_sheet.cell(row=new_row, column=1, value=f"{preferred_name} {surname}")
                    student_sheet.cell(row=new_row, column=2, value=year)  # Correctly set the year from column G
                    student_sheet.cell(row=new_row, column=3, value=form)
                    student_sheet.cell(row=new_row, column=4, value=school)
                    student_sheet.cell(row=new_row, column=5, value=instrument_id)
                    student_sheet.cell(row=new_row, column=6, value=date_of_loan_formatted)
                    student_sheet.cell(row=new_row, column=7, value=f"{duration_value} {duration_unit}")
                    student_sheet.cell(row=new_row, column=8, value=0)
                    student_sheet.cell(row=new_row, column=9, value=cost)
                    student_sheet.cell(row=new_row, column=10, value="")
                    student_sheet.cell(row=new_row, column=11, value=selected_student_id)
                    student_found = True
                    break
            if student_found:
                break

        if not student_found:
            print("Student not found in any division sheets.")

        try:
            wb.save('database.xlsx')
            print("Workbook saved successfully")
        except PermissionError:
            messagebox.showerror("Error", "Cannot add new items. Please close the database file (database.xlsx) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

        confirmation_label.config(text="Data submitted successfully!", fg="green")
        root.after(5000, clear_confirmation_label)
        
    def get_student_id(student_info):
        parts = student_info.split(" - ID: ")
        if len(parts) == 2:
            name_and_form, school_id = parts
            return school_id.strip()
        else:
            return ""

    def reposition_widgets(event):
        window_width = root.winfo_width()
        window_height = root.winfo_height()

        # Reposition the submit button
        button_width = submit_button.winfo_width()
        button_height = submit_button.winfo_height()
        x = (window_width - button_width) // 2
        y = window_height - button_height - 40
        submit_button.place(x=x, y=y)

        # Bring the back button to the front
        back_button.lift()

    def update_instrument_search_list(*args):
        search_terms = instrument_var.get().strip().lower().split()
        selected_school = school_var.get()
        search_listbox.delete(0, END)

        if search_terms:
            wb = load_workbook('database.xlsx')
            sheet = wb['Instruments']

            for row in sheet.iter_rows(min_row=2, values_only=True):
                instrument_name, instrument_id, description, serial_no, _, _, _, _, _, _, _, _, is_available, _, school = row[:15]
                if instrument_name and instrument_id:
                    instrument_info = f"{instrument_name} - ID: {instrument_id}"
                    match = True
                    for term in search_terms:
                        if not (
                            term in str(instrument_name).lower()
                            or term in str(instrument_id).lower()
                            or term in str(serial_no).lower()
                            or term in str(description).lower()
                        ):
                            match = False
                            break
                    if match and (selected_school == "ALL" or school == selected_school):
                        if is_available == 1:
                            search_listbox.insert(END, instrument_info)
                            search_listbox.itemconfig(END, foreground='green')
                        else:
                            search_listbox.insert(END, instrument_info)
                            search_listbox.itemconfig(END, foreground='red')

    def select_instrument(event):
        if search_listbox.curselection():
            selected_instrument = search_listbox.get(search_listbox.curselection()[0])
            instrument_name, instrument_id = selected_instrument.split(" - ID: ")
            instrument_var.set(instrument_name)
            instrument_entry.delete(0, END)
            instrument_entry.insert(0, instrument_name + " - ID: " + instrument_id)
            search_listbox.lower()

    def sort_key(forename, preferred_name, surname, form, tutor_email, school_id, year):
        if year and form:
            form_match = re.match(r"([LU]?)(\d+)([a-zA-Z]?)", form)
            if form_match:
                letter1, year_form, letter2 = form_match.groups()
                year_form = int(year_form)
                if not letter1:
                    letter1 = "L"
                if not letter2:
                    letter2 = "a"
                return int(year), (letter1 == "L", year_form, letter1, letter2.lower())
            else:
                return int(year), (True, float('inf'), "", "")
        else:
            return float('inf'), (True, float('inf'), "", "")

    def update_student_search_list(*args):
        search_terms = student_var.get().strip().lower().split()
        student_listbox.delete(0, END)

        if search_terms:
            wb = load_workbook('database.xlsx')
            sheets = ['Divisions_spj', 'Divisions_sps']
            student_info_list = []

            for sheet_name in sheets:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = list(row[:7])
                    forename, preferred_name, surname, form, tutor_email, school_id, year = row_data
                    student_info = f"{preferred_name} {surname} ({form}) - ID: {school_id}"
                    match = True
                    for term in search_terms:
                        if not (
                            term in str(forename).lower()
                            or term in str(preferred_name).lower()
                            or term in str(surname).lower()
                            or term in str(school_id).lower()
                        ):
                            match = False
                            break
                    if match:
                        student_info_list.append((student_info, *row_data))

            # Sort the student_info_list by year and then by form
            student_info_list.sort(key=lambda x: sort_key(*x[1:]))

            for data in student_info_list:
                student_info = data[0]
                student_listbox.insert(END, student_info)
                for term in search_terms:
                    if term.lower() in str(student_info).lower():
                        break
                else:
                    student_listbox.itemconfig(END)

    def show_date_picker():
        date_picker = Calendar(root, date_pattern="y-mm-dd")
        selected_date = date_picker.selection_get()
        if selected_date:
            date_of_loan_entry.delete(0, END)
            date_of_loan_entry.insert(0, selected_date)
        date_picker.destroy()

    def select_student(event):
        global selected_student_id
        if student_listbox.curselection():
            selected_student = student_listbox.get(student_listbox.curselection())
            student_id = get_student_id(selected_student)
            selected_student_id = student_id
            student_var.set(selected_student.split(" - ID: ")[0])
            student_listbox.lower()
            student_entry.config(width=len(student_var.get()))

    global confirmation_label
    confirmation_label = Label(root, bg="#E2DEDD", font=("Arial", 14))
    confirmation_label.pack(pady=(0, 20))

    input_frame = Frame(root, bg="#E2DEDD")
    input_frame.pack(pady=20, padx=20)

    # Column 1
    col1 = Frame(input_frame, bg="#E2DEDD")
    col1.pack(side=LEFT, padx=20, anchor=N)  # Anchor to the top

    student_name_label = Label(col1, text="Student Name:", bg="#E2DEDD", font=("Arial", 18))
    student_name_label.pack(pady=(0, 5))
    student_var = StringVar()
    student_entry = Entry(col1, textvariable=student_var, font=("Arial", 18), justify='center')
    student_entry.pack()

    student_listbox_frame = Frame(col1)
    student_listbox_frame.pack(pady=10, fill=BOTH, expand=True)

    student_listbox = Listbox(student_listbox_frame, font=("Arial", 14), height=5)
    student_listbox.pack(side=LEFT, fill=BOTH, expand=True)

    student_scrollbar = Scrollbar(student_listbox_frame)
    student_scrollbar.pack(side=RIGHT, fill=Y)
    student_listbox.config(yscrollcommand=student_scrollbar.set)
    student_scrollbar.config(command=student_listbox.yview)

    # Bind the select_student function after defining it
    student_listbox.bind("<<ListboxSelect>>", select_student)

    date_of_loan_label = Label(col1, text="Date of Loan: *", bg="#E2DEDD", font=("Arial", 18))
    date_of_loan_label.pack(pady=(20, 5))
    date_of_loan_entry = DateEntry(col1, font=("Arial", 18), justify='center', date_pattern='dd/MM/yyyy')
    date_of_loan_entry.pack(pady=5)

    date_picker_button = Button(col1, text="Select Date", command=show_date_picker)
    date_picker_button.pack(pady=10)

    # Column 2
    col2 = Frame(input_frame, bg="#E2DEDD")
    col2.pack(side=LEFT, padx=20, anchor=N)  # Anchor to the top

    duration_of_loan_label = Label(col2, text="Duration of Loan:", bg="#E2DEDD", font=("Arial", 18))
    duration_of_loan_label.pack(pady=(0, 5))
    duration_of_loan_frame = Frame(col2)
    duration_of_loan_frame.pack()
    duration_of_loan_entry = Entry(duration_of_loan_frame, font=("Arial", 18), justify='right', width=5)
    duration_of_loan_entry.pack(side=LEFT, padx=(0, 10))
    duration_of_loan_entry.insert(0, "1")  # Set the initial value to 1

    duration_units = ["Half Terms", "Terms", "Years"]
    duration_unit_var = StringVar(root)
    duration_unit_var.set(duration_units[0])  # Default option is "Terms"
    duration_unit_dropdown = ttk.Combobox(duration_of_loan_frame, textvariable=duration_unit_var, values=duration_units, state="readonly", width=10)
    duration_unit_dropdown.pack(side=LEFT)

    cost_label = Label(col2, text="Cost:", bg="#E2DEDD", font=("Arial", 18))
    cost_label.pack(pady=(20, 5))
    cost_entry_frame = Frame(col2)
    cost_entry_frame.pack()
    cost_currency_label = Label(cost_entry_frame, text="£", bg="#E2DEDD", font=("Arial", 18))
    cost_currency_label.pack(side=LEFT)
    cost_entry = Entry(cost_entry_frame, font=("Arial", 18), justify='right')
    cost_entry.pack(side=LEFT, fill=X, expand=True)
    cost_entry.insert(0, "0.00")  # Set the initial placeholder

    instrument_label = Label(col2, text="Instrument:", bg="#E2DEDD", font=("Arial", 18))
    instrument_label.pack(pady=(20, 5))
    instrument_var = StringVar()
    instrument_entry = Entry(col2, textvariable=instrument_var, font=("Arial", 18))
    instrument_entry.pack()

    # Create a frame to hold the search listbox and legend
    search_legend_frame = Frame(col2, bg="#E2DEDD")
    search_legend_frame.pack(pady=10)

    search_listbox = Listbox(search_legend_frame, font=("Arial", 14), height=5)
    search_listbox.pack(side=LEFT, padx=(0, 20))

    student_var.trace_add("write", update_student_search_list)

    search_listbox.bind("<<ListboxSelect>>", select_instrument)

    instrument_var.trace_add("write", update_instrument_search_list)



    # Submit Button
    submit_button = Button(root, text="Submit", font=("Arial", 18), command=submit_data_out)
    submit_button.pack(pady=20)

    def reposition_button_frame(event):
        window_width = root.winfo_width()
        button_frame_width = button_frame.winfo_reqwidth()
        x = (window_width - button_frame_width) // 2
        y = 10  # Adjust the y-coordinate as needed
        button_frame.place(x=x, y=y)

    button_frame = Frame(root, bg="#E2DEDD")

    school_button = Button(button_frame, textvariable=school_var, command=toggle_school, font=("Arial", 14))
    school_button.pack(side=TOP)

    reposition_button_frame(None)  # Call the function to position the button frame initially

    root.bind("<Configure>", reposition_button_frame)

    root.bind("<Configure>", reposition_widgets)

def show_info_interface(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_info_interface)

    back_button = create_back_button(root, lambda: homepage(root))

    instrument_info_button = Button(root, text="Instrument Information", bg="purple", fg="white", font=("Arial", 24), command=lambda: show_instrument_info(root))
    student_info_button = Button(root, text="Student Information", bg="orange", fg="white", font=("Arial", 24), command=lambda: show_student_info(root))

    buttons = [instrument_info_button, student_info_button]
    root.bind("<Configure>", lambda event: resize_info_buttons(root, buttons))
    resize_info_buttons(root, buttons)  # Initial button placement

    def raise_back_button(event):
        back_button.lift()

    root.bind("<Configure>", raise_back_button)

def show_modify_interface(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_modify_interface)

    back_button = create_back_button(root, lambda: homepage(root))

    student_modify_button = Button(root, text="Modify Student", bg="purple", fg="white", font=("Arial", 24), command=lambda: modify_student(root))
    loans_modify_button = Button(root, text="Modify Current Loans", bg="orange", fg="white", font=("Arial", 24), command=lambda: modify_loans(root))
    instruments_modify_button = Button(root, text="Modify Instruments", bg="blue", fg="white", font=("Arial", 24), command=lambda: show_instrument_search(root))

    buttons = [student_modify_button, loans_modify_button, instruments_modify_button]
    root.bind("<Configure>", lambda event: resize_modify_buttons(root, buttons))
    resize_modify_buttons(root, buttons)  # Initial button placement

    def raise_back_button(event):
        back_button.lift()

    root.bind("<Configure>", raise_back_button)

def resize_modify_buttons(root, buttons):
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    padding = 20
    button_width = window_width - padding * 2
    button_height = (window_height - padding * 5) // 3

    buttons[0].place(x=padding, y=padding, width=button_width, height=button_height)
    buttons[1].place(x=padding, y=padding * 2 + button_height, width=button_width, height=button_height)
    buttons[2].place(x=padding, y=padding * 3 + button_height * 2, width=button_width, height=button_height)

def modify_student(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(modify_student)

    back_button = create_back_button(root, lambda: homepage(root))
    # Add code for modifying student data here

def modify_loans(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(modify_loans)

    back_button = create_back_button(root, lambda: homepage(root))
    # Add code for modifying current  loans data here

def resize_instrument_buttons(root, buttons):
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    padding = 10
    button_width = window_width - padding * 2
    button_height = (window_height - padding * (len(buttons) + 1)) // len(buttons)

    for i, button in enumerate(buttons):
        button.place(x=padding, y=padding * (i + 1) + button_height * i, width=button_width, height=button_height)

def show_instrument_search(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_instrument_search)
    back_button = create_back_button(root, lambda: homepage(root))


    # Search bar
    search_frame = Frame(root, bg="#E2DEDD")
    search_frame.pack(pady=10)

    search_label = Label(search_frame, text="Search:", bg="#E2DEDD", font=("Arial", 14))
    search_label.pack(side=LEFT, padx=5)

    search_entry = Entry(search_frame, font=("Arial", 14))
    search_entry.pack(side=LEFT, padx=5)

    search_button = Button(search_frame, text="Search", font=("Arial", 14), command=lambda: search_instruments(search_entry.get()))
    search_button.pack(side=LEFT, padx=5)

    # Result area
    result_area = Frame(root)
    result_area.pack(pady=10, padx=10, fill=BOTH, expand=True)

    # Create a treeview with column titles
    result_tree = ttk.Treeview(result_area, columns=("Name", "ID", "Serial No", "School"), show="headings")
    result_tree.heading("Name", text="Name")
    result_tree.heading("ID", text="ID")
    result_tree.heading("Serial No", text="Serial No")
    result_tree.heading("School", text="School")
    # Adjust column widths
    result_tree.column("Name", width=250)
    result_tree.column("ID", width=100)
    result_tree.column("Serial No", width=150)
    result_tree.column("School", width=100)

    # Add the treeview to the result area
    result_tree.pack(side=LEFT, fill=BOTH, expand=True)

    # Add a vertical scrollbar to the treeview
    vsb = ttk.Scrollbar(result_area, orient="vertical", command=result_tree.yview)
    vsb.pack(side=RIGHT, fill=Y)
    result_tree.configure(yscrollcommand=vsb.set)

    def search_instruments(query):
        # Load the workbook
        wb = load_workbook('database.xlsx')
        sheet = wb['Instruments']

        # Clear the result tree
        result_tree.delete(*result_tree.get_children())

        # Search for matching instruments
        for row in sheet.iter_rows(min_row=2, values_only=True):
            instrument_name, instrument_id, description, serial_no, _, _, _, _, _, _, _, _, is_available = row[:13]
            if query.lower() in str(instrument_name).lower() or query.lower() in str(instrument_id).lower() or query.lower() in str(serial_no).lower():
                if str(instrument_name) != "None" or str(instrument_id) != "None" or str(serial_no) != "None":
                    available_text = "Yes" if is_available == 1 else "No"
                    school = row[14]
                    tag = "available" if is_available == 1 else "unavailable"
                    result_tree.insert("", END, values=(instrument_name, instrument_id, serial_no, school), tags=(tag,))

        # Configure tags for color-coding
        result_tree.tag_configure("available", background="#50AF54")
        result_tree.tag_configure("unavailable", background="red")

    def modify_selected_instrument(event):
        selected_item = result_tree.focus()
        if selected_item:
            item_values = result_tree.item(selected_item)["values"]
            instrument_name, instrument_id, serial_no, _ = item_values
            modify_instrument_info(root, instrument_name, instrument_id, serial_no)
        else:
            back_button.lift()  # Bring the back button to the front

    # Bind the double-click event to the modify_selected_instrument function
    result_tree.bind("<Double-1>", modify_selected_instrument)

    # Bring the back button to the front
    back_button.lift()

def modify_instrument_info(root, instrument_name, instrument_id, serial_no):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(modify_instrument_info)
    back_button = create_back_button(root, lambda: homepage(root))

    fields = ["Instrument Name", "Instrument Description", "Serial Number", "Date of Last Valuation",
              "Last valuation amount", "Last repair date", "Purchased date", "Vendor", "Notes",
              "Maintenance Notes", "Is Available", "Storage location", "School", "Category"]

    buttons = []
    for field in fields:
        button = Button(root, text=f"Modify {field}", bg="purple", fg="white", font=("Arial", 16),
                        command=lambda field_name=field: modify_instrument_field(root, instrument_name, instrument_id, serial_no, field_name))
        buttons.append(button)

    root.bind("<Configure>", lambda event: resize_instrument_buttons(root, buttons))
    resize_instrument_buttons(root, buttons)  # Initial button placement

    # Bring the back button to the front
    back_button.lift()
    
def modify_instrument_field(root, instrument_name, instrument_id, serial_no, field_name):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(modify_instrument_field)
    back_button = create_back_button(root, lambda: homepage(root))

    # Create a label to display the instrument details
    instrument_details_label = Label(root, text=f"Instrument: {instrument_name} - ID: {instrument_id} - Serial No: {serial_no}", font=("Arial", 18, "bold"), bg="#E2DEDD")
    instrument_details_label.pack(pady=10)

    # Load the workbook
    wb = load_workbook('database.xlsx')
    sheet = wb['Instruments']

    # Find the row with the matching instrument details
    row_index = None
    for row in range(2, sheet.max_row + 1):
        row_instrument_id = sheet.cell(row=row, column=2).value
        row_instrument_name = sheet.cell(row=row, column=1).value

        print(f"Checking row: Instrument ID = '{row_instrument_id}', Instrument Name = '{row_instrument_name}'")

        if str(row_instrument_id) == str(instrument_id) and str(row_instrument_name) == str(instrument_name):
            row_index = row
            break

    if row_index is None:
        print("Instrument not found in the database.")
        messagebox.showerror("Error", "Instrument not found.")
        back_button.lift()  # Bring the back button to the front
        return

    print(f"Instrument found in row {row_index}")

    # Get the column index for the specified field
    column_index = None
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col).value).lower() == field_name.lower():
            column_index = col
            break


    if column_index is None:
        print(f"Column '{field_name}' not found.")
        messagebox.showerror("Error", f"Column '{field_name}' not found.")
        back_button.lift()  # Bring the back button to the front
        return

    # Get the current value of the field
    current_value = sheet.cell(row=row_index, column=column_index).value

    # Create a label for the field
    label = Label(root, text=f"Current {field_name}:", font=("Arial", 14), bg="#E2DEDD")
    label.pack(pady=10)

    datetime_fields = ["date of last valuation", "last repair date", "purchased date"]
    if field_name.lower() in datetime_fields:
        # Use a clickable DateEntry for datetime fields
        entry = DateEntry(root, font=("Arial", 14), date_pattern='dd/MM/yyyy')
        if current_value:
            entry.set_date(current_value.date())
        entry.pack(pady=10)
    elif field_name.lower() == "category":
        # Create a dropdown for the "Category" field
        category_options = ["String", "Woodwind", "Brass", "Percussion", "Keyboard", "Other"]
        entry = ttk.Combobox(root, values=category_options, font=("Arial", 14))
        entry.set(current_value)  # Set the current value
        entry.pack(pady=10)
    elif field_name.lower() == "school":
        # Create a dropdown for the "Category" field
        category_options = ["JUNIORS", "SENIORS"]
        entry = ttk.Combobox(root, values=category_options, font=("Arial", 14))
        entry.set(current_value)  # Set the current value
        entry.pack(pady=10)
    elif field_name.lower() == "is available":
        # Create a dropdown for the "Is Available" field
        availability_options = ["True", "False"]
        entry = ttk.Combobox(root, values=availability_options, font=("Arial", 14))
        entry.set(str(bool(current_value)))  # Set the current value
        entry.pack(pady=10)
    else:
        entry = Entry(root, font=("Arial", 14))
        entry.insert(0, str(current_value))
        entry.pack(pady=10)

    def save_changes():
        if field_name.lower() in datetime_fields:
            new_value = entry.get_date().strftime('%d/%m/%Y')
        elif field_name.lower() == "is available":
            new_value = 1 if entry.get() == "True" else 0
        else:
            new_value = entry.get()

        sheet.cell(row=row_index, column=column_index, value=new_value)
        wb.save('database.xlsx')

        # Create a confirmation label
        confirmation_label = Label(root, text="Item added successfully! ✓", fg="green", bg="#E2DEDD", font=("Arial", 14))
        confirmation_label.pack(pady=10)

        modify_instrument_info(root, instrument_name, instrument_id, serial_no)

        # Remove the confirmation label after 5 seconds
        root.after(5000, confirmation_label.pack_forget)


    save_button = Button(root, text="Save Changes", font=("Arial", 14), command=save_changes)
    save_button.pack(pady=10)

    text_placeholder = Label(text="test", bg="#E2DEDD", font=("Arial", 18))

    # Bring the back button to the front
    back_button.lift()

def resize_info_buttons(root, buttons):
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    padding = 20
    button_width = window_width - padding * 2
    button_height = (window_height - padding * 4) // 2

    buttons[0].place(x=padding, y=padding, width=button_width, height=button_height)
    buttons[1].place(x=padding, y=padding + button_height + padding, width=button_width, height=button_height)

def show_instrument_info(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_instrument_info)
    back_button = create_back_button(root, lambda: homepage(root))

    # Search bar
    search_frame = Frame(root, bg="#E2DEDD")
    search_frame.pack(pady=10)

    search_label = Label(search_frame, text="Search:", bg="#E2DEDD", font=("Arial", 14))
    search_label.pack(side=LEFT, padx=5)

    search_entry = Entry(search_frame, font=("Arial", 14))
    search_entry.pack(side=LEFT, padx=5)

    search_button = Button(search_frame, text="Search", font=("Arial", 14), command=lambda: search_instruments(search_entry.get()))
    search_button.pack(side=LEFT, padx=5)

    # Result area
    result_area = Frame(root)
    result_area.pack(pady=10, padx=10, fill=BOTH, expand=True)

    # Create a treeview with column titles
    result_tree = ttk.Treeview(result_area, columns=("Name", "ID", "Serial No", "School"), show="headings")
    result_tree.heading("Name", text="Name")
    result_tree.heading("ID", text="ID")
    result_tree.heading("Serial No", text="Serial No")
    result_tree.heading("School", text="School")

    # Adjust column widths
    result_tree.column("Name", width=250)
    result_tree.column("ID", width=100)
    result_tree.column("Serial No", width=150)
    result_tree.column("School", width=100)

    # Add the treeview to the result area
    result_tree.pack(side=LEFT, fill=BOTH, expand=True)

    # Add a vertical scrollbar to the treeview
    vsb = ttk.Scrollbar(result_area, orient="vertical", command=result_tree.yview)
    vsb.pack(side=RIGHT, fill=Y)
    result_tree.configure(yscrollcommand=vsb.set)

    def search_instruments(query):
        # Load the workbook
        wb = load_workbook('database.xlsx')
        sheet = wb['Instruments']

        # Clear the result tree
        result_tree.delete(*result_tree.get_children())

        # Search for matching instruments
        for row in sheet.iter_rows(min_row=2, values_only=True):
            instrument_name, instrument_id, description, serial_no, _, _, _, _, _, _, _, _, is_available, _, school = row[:15]
            if query.lower() in str(instrument_name).lower() or query.lower() in str(instrument_id).lower() or query.lower() in str(serial_no).lower():
                if str(instrument_name) != "None" or str(instrument_id) != "None" or str(serial_no) != "None":
                    tag = "available" if is_available == 1 else "unavailable"
                    result_tree.insert("", END, values=(instrument_name, instrument_id, serial_no, school), tags=(tag,))

        # Configure tags for color-coding
        result_tree.tag_configure("available", background="#50AF54")
        result_tree.tag_configure("unavailable", background="red")


    def show_instrument_details(event):
        selected_item = result_tree.focus()
        if selected_item:
            item_values = result_tree.item(selected_item)["values"]
            instrument_name, instrument_id, serial_no, school = [str(value) for value in item_values]

            # Load the workbook
            wb = load_workbook('database.xlsx')
            sheet = wb['Instruments']

            # Find the row containing the selected instrument details
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_instrument_name, row_instrument_id, _, row_serial_no, _, _, _, _, _, _, _, _, is_available, _, row_school = [str(value) for value in row[:15]]
                if row_instrument_name.strip() == instrument_name.strip() and row_instrument_id.strip() == instrument_id.strip() and row_serial_no.strip() == serial_no.strip():
                    instrument_details = row

                    # Create a new popup window
                    details_window = tk.Toplevel(root)
                    details_window.title("Instrument Details")

                    # Add labels and values to the popup window
                    row_num = 0
                    for column_num, value in enumerate(instrument_details):
                        column_name = sheet.cell(row=1, column=column_num + 1).value
                        if column_name == "Is Available":
                            value = "Yes" if value == 1 else "No"
                        label = tk.Label(details_window, text=f"{column_name}: {value}", font=("Arial", 12))
                        label.grid(row=row_num, column=0, sticky="w", padx=10, pady=5)
                        row_num += 1

                    break
            else:
                print("Instrument details not found in the database.")

    legend_frame = Frame(root, bg="#E2DEDD")
    legend_frame.pack(pady=10)

    available_label = Label(legend_frame, text="Available", bg="#50AF54", fg="white", font=("Arial", 12), padx=5, pady=2)
    available_label.pack(side=LEFT, padx=5)

    unavailable_label = Label(legend_frame, text="Unavailable", bg="red", fg="white", font=("Arial", 12), padx=5, pady=2)
    unavailable_label.pack(side=LEFT, padx=5)

    # Bind the double-click event to the show_instrument_details function
    result_tree.bind("<Double-1>", show_instrument_details)

def show_student_info(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_student_info)
    back_button = create_back_button(root, lambda: homepage(root))

    # Search bar
    search_frame = Frame(root, bg="#E2DEDD")
    search_frame.pack(pady=10)

    search_label = Label(search_frame, text="Search:", bg="#E2DEDD", font=("Arial", 14))
    search_label.pack(side=LEFT, padx=5)

    search_entry = Entry(search_frame, font=("Arial", 14))
    search_entry.pack(side=LEFT, padx=5)

    search_button = Button(search_frame, text="Search", font=("Arial", 14), command=lambda: search_students(search_entry.get()))
    search_button.pack(side=LEFT, padx=5)

    loan_histories_button = Button(search_frame, text="Loan Histories", font=("Arial", 14), command=lambda: show_history_interface(root))
    loan_histories_button.pack(side=LEFT, padx=5)

    # Result area
    result_area = Frame(root)
    result_area.pack(pady=10, padx=10, fill=BOTH, expand=True)

    # Create a treeview with column titles
    result_tree = ttk.Treeview(result_area, columns=("Name", "Form", "Currently Hiring", "Date of Return"), show="headings")
    result_tree.heading("Name", text="Name")
    result_tree.heading("Form", text="Form")
    result_tree.heading("Currently Hiring", text="Currently Hiring")
    result_tree.heading("Date of Return", text="Date of Return")

    # Adjust column widths
    result_tree.column("Name", width=200)
    result_tree.column("Form", width=100)
    result_tree.column("Currently Hiring", width=200)
    result_tree.column("Date of Return", width=150)

    # Add the treeview to the result area
    result_tree.pack(side=LEFT, fill=BOTH, expand=True)

    # Add a vertical scrollbar to the treeview
    vsb = ttk.Scrollbar(result_area, orient="vertical", command=result_tree.yview)
    vsb.pack(side=RIGHT, fill=Y)
    result_tree.configure(yscrollcommand=vsb.set)

    def search_students(query=""):
        try:
            # Load the workbook
            wb = load_workbook('database.xlsx')
            student_sheet = wb['Student']
            instruments_sheet = wb['Instruments']
            current_loans_sheet = wb['Current_Loans']

            # Clear the result tree
            result_tree.delete(*result_tree.get_children())

            # Get the current date
            today = date.today()

            # Search for matching students
            for student_row in student_sheet.iter_rows(min_row=2, values_only=True):
                name, year, form, school, instrument_id, date_of_hire, duration, form_signed, charge, returned, student_id = (list(student_row) + [None] * 11)[:11]

                # Convert date_of_hire to date object if it's a string
                if isinstance(date_of_hire, str):
                    date_of_hire = datetime.strptime(date_of_hire, "%d%m%Y").date()

                # Convert date_of_hire to a string for searching
                if date_of_hire:
                    date_of_hire_str = date_of_hire.strftime("%d/%m/%Y")
                else:
                    date_of_hire_str = ""

                # Check if the student matches the search query
                if (
                    query.lower() in str(name).lower()
                    or query.lower() in str(school).lower()
                    or query.lower() in str(instrument_id).lower()
                    or query.lower() in date_of_hire_str.lower()
                    or query.lower() in str(duration).lower()
                ):
                    # Find the instrument details for the student
                    for instrument_row in instruments_sheet.iter_rows(min_row=2, values_only=True):
                        instrument_row_id, instrument_name, *_ = instrument_row
                        if instrument_row_id == instrument_id:
                            instrument_info = f"{instrument_name} - ID: {instrument_id}"
                            break
                    else:
                        instrument_info = f"ID: {instrument_id}"

                    # Get the actual date of return from the Current_Loans sheet
                    for loan_row in current_loans_sheet.iter_rows(min_row=2, values_only=True):
                        loan_student_id, _, _, _, _, date_of_return, _ = loan_row
                        if str(student_id) == str(loan_student_id):
                            actual_date_of_return = date_of_return if date_of_return else "NONE"
                            break
                    else:
                        actual_date_of_return = "NONE"

                    print(actual_date_of_return)

                    values = (name, form, instrument_info, actual_date_of_return)
                    result_tree.insert("", END, values=values)

            # Print the search results to the console
            print("Search Results:")
            for item in result_tree.get_children():
                values = result_tree.item(item)['values']
                print(f"Name: {values[0]}, Form: {values[1]}, Instrument: {values[2]}, Date of Return: {values[3]}")

        except PermissionError:
            messagebox.showerror("Error", "Cannot access the database. Please close the database file (database.xlsx) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def show_loan_details(event):
        try:
            selected_item = result_tree.focus()
            if selected_item:
                item_values = result_tree.item(selected_item)["values"]

                # Check if item_values has at least 2 elements
                if len(item_values) >= 2:
                    student_name = item_values[0]  # Student Name
                    student_form = item_values[1]  # Form

                    # Load the workbook
                    wb = load_workbook('database.xlsx')
                    student_sheet = wb['Student']
                    loans_sheet = wb['Current_Loans']
                    instruments_sheet = wb['Instruments']

                    # Find the student ID based on the student name and form
                    for row in student_sheet.iter_rows(min_row=2, values_only=True):
                        name, _, form, _, _, _, _, _, _, _, student_id = row
                        if name == student_name and form == student_form:
                            break
                    else:
                        # Student not found
                        return

                    # Find the loan details for the selected student
                    for loan_row in loans_sheet.iter_rows(min_row=2, values_only=True):
                        loan_student_id, instrument_id, date_of_loan, duration, cost, date_of_return, returned = loan_row[:7]

                        if str(student_id) == str(loan_student_id) and returned != 1:
                            # Find the instrument name from the "Instruments" sheet
                            for instrument_row in instruments_sheet.iter_rows(min_row=2, values_only=True):
                                _, instrument_row_id, instrument_name, *_ = instrument_row
                                if str(instrument_id) == str(instrument_row_id):
                                    instrument_info = f"{instrument_name} - ID: {instrument_id}"
                                    break
                            else:
                                instrument_info = f"ID: {instrument_id}"

                            # Create a new popup window
                            details_window = tk.Toplevel(root)
                            details_window.title("Loan Details")

                            # Add labels and values to the popup window
                            tk.Label(details_window, text=f"Instrument: {instrument_info}", font=("Arial", 12)).pack(pady=5)
                            tk.Label(details_window, text=f"Date of Loan: {date_of_loan}", font=("Arial", 12)).pack(pady=5)
                            tk.Label(details_window, text=f"Duration: {duration}", font=("Arial", 12)).pack(pady=5)
                            tk.Label(details_window, text=f"Cost: {cost}", font=("Arial", 12)).pack(pady=5)
                            tk.Label(details_window, text=f"Date of Return: {date_of_return}", font=("Arial", 12)).pack(pady=5)

                            break
                else:
                    # Handle the case where item_values doesn't have enough elements
                    print("Error: item_values doesn't have enough elements")

        except Exception as e:
            # Print the error traceback to the console
            print(f"An error occurred: {e}")
            traceback.print_exc()

            # Find the loan details for the selected student
            for loan_row in loans_sheet.iter_rows(min_row=2, values_only=True):
                loan_student_id, instrument_id, date_of_loan, duration, cost, date_of_return, returned = loan_row

                if str(student_id) == str(loan_student_id) and returned != 1:
                    # Find the instrument name from the "Instruments" sheet
                    for instrument_row in instruments_sheet.iter_rows(min_row=2, values_only=True):
                        _, instrument_row_id, instrument_name, *_ = instrument_row
                        if str(instrument_id) == str(instrument_row_id):
                            instrument_info = f"{instrument_name} - ID: {instrument_id}"
                            break
                    else:
                        instrument_info = f"ID: {instrument_id}"

                    # Create a new popup window
                    details_window = tk.Toplevel(root)
                    details_window.title("Loan Details")

                    # Add labels and values to the popup window
                    tk.Label(details_window, text=f"Instrument: {instrument_info}", font=("Arial", 12)).pack(pady=5)
                    tk.Label(details_window, text=f"Date of Loan: {date_of_loan}", font=("Arial", 12)).pack(pady=5)
                    tk.Label(details_window, text=f"Duration: {duration}", font=("Arial", 12)).pack(pady=5)
                    tk.Label(details_window, text=f"Cost: {cost}", font=("Arial", 12)).pack(pady=5)
                    tk.Label(details_window, text=f"Date of Return: {date_of_return}", font=("Arial", 12)).pack(pady=5)

                    break
    # Bind the double-click event to the show_loan_details function
    result_tree.bind("<Double-1>", show_loan_details)

    search_button.invoke()  # Call the search_students function initial

def show_history_interface(root):    
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_history_interface)
    back_button = create_back_button(root, lambda: homepage(root))


    # Search bar
    search_frame = Frame(root, bg="#E2DEDD")
    search_frame.pack(pady=10)

    search_label = Label(search_frame, text="Search:", bg="#E2DEDD", font=("Arial", 14))
    search_label.pack(side=LEFT, padx=5)

    search_entry = Entry(search_frame, font=("Arial", 14))
    search_entry.pack(side=LEFT, padx=5)

    search_button = Button(search_frame, text="Search", font=("Arial", 14), command=lambda: search_loan_histories(search_entry.get()))
    search_button.pack(side=LEFT, padx=5)

    # Dropdown menus
    dropdown_frame = Frame(root, bg="#E2DEDD")
    dropdown_frame.pack(pady=10)

    def cycle_school():
        schools = ["ALL", "SENIORS", "JUNIORS"]
        current_index = schools.index(school_var.get())
        next_index = (current_index + 1) % len(schools)
        school_var.set(schools[next_index])
        update_form_dropdown()
        search_loan_histories(search_entry.get())

    # School button
    school_var = StringVar(value="ALL")
    school_button = Button(dropdown_frame, textvariable=school_var, font=("Arial", 14), command=cycle_school)
    school_button.pack(side=LEFT, padx=5)

    # Dropdown menu for form
    form_var = StringVar(value="None")
    form_dropdown = OptionMenu(dropdown_frame, form_var, "None")
    form_dropdown.config(font=("Arial", 14))
    form_dropdown.pack(side=LEFT, padx=5)

    # Result area
    result_area = Frame(root)
    result_area.pack(pady=10, padx=10, fill=BOTH, expand=True)

    # Create a treeview with column titles
    result_tree = ttk.Treeview(result_area, columns=("Student Name", "School", "Form", "Instrument", "Date of Loan", "Date of Return"), show="headings")
    result_tree.heading("Student Name", text="Student Name")
    result_tree.heading("School", text="School")
    result_tree.heading("Form", text="Form")
    result_tree.heading("Instrument", text="Instrument")
    result_tree.heading("Date of Loan", text="Date of Loan")
    result_tree.heading("Date of Return", text="Date of Return")

    # Adjust column widths
    result_tree.column("Student Name", width=200)
    result_tree.column("School", width=100)
    result_tree.column("Form", width=100)
    result_tree.column("Instrument", width=200)
    result_tree.column("Date of Loan", width=150)
    result_tree.column("Date of Return", width=150)

    # Add the treeview to the result area
    result_tree.pack(side=LEFT, fill=BOTH, expand=True)

    # Add a vertical scrollbar to the treeview
    vsb = ttk.Scrollbar(result_area, orient="vertical", command=result_tree.yview)
    vsb.pack(side=RIGHT, fill=Y)
    result_tree.configure(yscrollcommand=vsb.set)
    
    def search_loan_histories(query=""):
        # Load the workbook
        try:
            wb = load_workbook('database.xlsx')
        except PermissionError:
            messagebox.showerror("Error", "Cannot access the database. Please close the database file (database.xlsx) and try again.")
            return
        student_sheet = wb['Student']
        instruments_sheet = wb['Instruments']
        current_loans_sheet = wb['Current_Loans']

        # Clear the result tree
        result_tree.delete(*result_tree.get_children())

        # Get the selected school from the dropdown menu
        selected_school = school_var.get()
        selected_form = form_var.get()

        # Create a dictionary to map student IDs to their details
        student_dict = {}
        for student_row in student_sheet.iter_rows(min_row=2, values_only=True):
            name = student_row[0]  # Column A
            form = student_row[2]  # Column C
            school = student_row[3]  # Column D
            instrument_id = student_row[4]  # Column E
            student_id = student_row[10]  # Column K
            student_dict[str(student_id)] = (name, school, form, instrument_id)

        # Create a dictionary to map instrument IDs to instrument names
        instrument_dict = {}
        for instrument_row in instruments_sheet.iter_rows(min_row=2, values_only=True):
            instrument_id = instrument_row[1]  # Assuming the ID is in the second column
            instrument_name = instrument_row[0]  # Assuming the name is in the first column
            instrument_dict[str(instrument_id)] = instrument_name

        # Search for matching loans
        for loan_row in current_loans_sheet.iter_rows(min_row=2, values_only=True):
            pupil_id, instrument_id, date_of_loan, duration, cost, date_of_return, returned = loan_row

            # Find the student details for the loan
            student_details = student_dict.get(str(pupil_id))
            if student_details:
                name, school, form, student_instrument_id = student_details

                # Convert date_of_loan to the desired string format
                date_of_loan_str = date_of_loan[:2] + "/" + date_of_loan[2:4] + "/" + date_of_loan[4:]

                # Convert date_of_return to the desired string format if it exists
                if date_of_return:
                    date_of_return_str = date_of_return[:2] + "/" + date_of_return[2:4] + "/" + date_of_return[4:]
                else:
                    date_of_return_str = "NONE"

                # Check if the student matches the selected school, form, and the search query
                if (
                    (selected_school == "ALL" or selected_school == "NONE" or (selected_school == "SENIORS" and school == "SPS") or (selected_school == "JUNIORS" and school == "SPJ"))
                    and (selected_form == "None" or form == selected_form)
                    and (
                        query.lower() in str(name).lower()
                        or query.lower() in str(form).lower()
                        or query.lower() in str(instrument_id).lower()
                        or query.lower() in date_of_loan_str.lower()
                        or query.lower() in instrument_dict.get(str(instrument_id), "").lower()
                    )
                ):
                    # Get the instrument name from the dictionary
                    instrument_name = instrument_dict.get(str(instrument_id), f"Unknown - ID: {instrument_id}")

                    # Determine the tag for highlighting
                    if date_of_return_str == "NONE":
                        tag = "none"
                    else:
                        date_of_return_date = datetime.strptime(date_of_return_str, "%d/%m/%Y").date()
                        if date_of_return_date < date.today():
                            tag = "past"
                        else:
                            tag = "future"

                    # Add the student's details to the result tree
                    result_tree.insert("", END, values=(name, school, form, instrument_name + ", ID: " + instrument_id, date_of_loan_str, date_of_return_str), tags=(tag,))

        # Configure tags for color-coding
        result_tree.tag_configure("none", background="yellow")
        result_tree.tag_configure("past", background="red")
        result_tree.tag_configure("future", background="green")




    def update_form_dropdown(*args):
        selected_school = school_var.get()

        # Clear the form dropdown
        form_dropdown['menu'].delete(0, 'end')

        # Load the workbook
        wb = load_workbook('database.xlsx')

        # Get the appropriate form options based on the selected school
        forms = set()
        if selected_school in ["ALL", "SENIORS"]:
            sheet = wb['Divisions_sps']
            forms.update(row[3] for row in sheet.iter_rows(min_row=2, values_only=True) if row[3])
        if selected_school in ["ALL", "JUNIORS"]:
            sheet = wb['Divisions_spj']
            forms.update(row[3] for row in sheet.iter_rows(min_row=2, values_only=True) if row[3])

        # Remove duplicates and sort the forms
        forms = sorted(forms, key=form_sorting_key)

        # Add the forms to the dropdown menu
        for form in forms:
            form_dropdown['menu'].add_command(label=form, command=lambda value=form: form_var.set(value))

        # Set the default value
        form_var.set("None")

    # Initial population of the form dropdown
    update_form_dropdown()

    # Bind the school_var to update the form dropdown when it changes
    school_var.trace('w', update_form_dropdown)

    # Bind the form_var to update the search results when it changes
    form_var.trace('w', lambda *args: search_loan_histories(search_entry.get()))

    # Search loan histories initially
    search_loan_histories()

def show_returns_interface(root):
    root.unbind("<<ClearWindow>>")
    clear_window(root)
    page_stack.append(show_returns_interface)

    back_button = create_back_button(root, lambda: go_back(root))

    # Search bar
    search_frame = Frame(root, bg="#E2DEDD")
    search_frame.pack(pady=10)

    search_label = Label(search_frame, text="Search:", bg="#E2DEDD", font=("Arial", 14))
    search_label.pack(side=LEFT, padx=5)

    search_entry = Entry(search_frame, font=("Arial", 14))
    search_entry.pack(side=LEFT, padx=5)

    search_button = Button(search_frame, text="Search", font=("Arial", 14), command=lambda: search_current_loans(search_entry.get()))
    search_button.pack(side=LEFT, padx=5)

    # Result area
    result_area = Frame(root)
    result_area.pack(pady=10, padx=10, fill=BOTH, expand=True)

    # Create a treeview with column titles
    result_tree = ttk.Treeview(result_area, columns=("Pupil", "Instrument", "Date of Loan", "Date of Return"), show="headings")
    result_tree.heading("Pupil", text="Pupil")
    result_tree.heading("Instrument", text="Instrument")
    result_tree.heading("Date of Loan", text="Date of Loan")
    result_tree.heading("Date of Return", text="Date of Return")

    # Adjust column widths
    result_tree.column("Pupil", width=200)
    result_tree.column("Instrument", width=200)
    result_tree.column("Date of Loan", width=150)
    result_tree.column("Date of Return", width=150)

    # Add the treeview to the result area
    result_tree.pack(side=LEFT, fill=BOTH, expand=True)

    # Add a vertical scrollbar to the treeview
    vsb = ttk.Scrollbar(result_area, orient="vertical", command=result_tree.yview)
    vsb.pack(side=RIGHT, fill=Y)
    result_tree.configure(yscrollcommand=vsb.set)
   
    def delete_selected_loan(event):
        selected_item = result_tree.selection()
        if selected_item:
            item_values = result_tree.item(selected_item)['values']
            if item_values:
                student_name, instrument_name_str, date_of_loan_str, date_of_return_str = item_values

                # Extract the instrument ID from the instrument_name_str
                instrument_id = instrument_name_str.split(": ")[-1]

                print(f"Selected loan values: {student_name}, {instrument_id}, {date_of_loan_str}, {date_of_return_str}")

                confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete the loan for {student_name} ({instrument_name_str})?")
                if confirm:
                    try:
                        # Load the workbook
                        wb = load_workbook('database.xlsx')
                        loans_sheet = wb['Current_Loans']
                        instruments_sheet = wb['Instruments']

                        # Find the row to delete in the Current_Loans sheet
                        row_to_delete = None
                        for row in loans_sheet.iter_rows(min_row=2, values_only=True):
                            _, id, _, _, _, _, _ = row
                            if id == instrument_id:
                                row_to_delete = row
                                break

                        print(f"row_to_delete: {row_to_delete}")

                        if row_to_delete:
                            # Delete the row from Current_Loans
                            for row_num, row in enumerate(loans_sheet.iter_rows(min_row=2, values_only=True), start=2):
                                if row == row_to_delete:
                                    loans_sheet.delete_rows(row_num, 1)
                                    break

                            # Change Is Available in Instruments to 1 and clear Currently Hired By
                            for row in instruments_sheet.iter_rows(min_row=2, values_only=True):
                                if str(row[1]) == str(instrument_id):
                                    print(f"Matching row found: {row}")
                                    instrument_row = row[0].row
                                    instruments_sheet.cell(row=instrument_row, column=13, value=1)  # Set column M (Is Available) to 1
                                    instruments_sheet.cell(row=instrument_row, column=12).value = None  # Clear column L (Currently Hired By)
                                    break

                            # Save the workbook
                            wb.save('database.xlsx')

                            # Remove the selected item from the treeview
                            result_tree.delete(selected_item)

                            messagebox.showinfo("Success", "Loan deleted successfully.")
                        else:
                            messagebox.showerror("Error", "No matching loan found in the database.")

                    except PermissionError:
                        messagebox.showerror("Error", "Cannot access the database. Please close the database file (database.xlsx) and try again.")
                    except Exception as e:
                        messagebox.showerror("Error", f"An error occurred: {str(e)}")
            else:
                messagebox.showerror("Error", "No item selected in the treeview.")
        else:
            messagebox.showerror("Error", "No item selected in the treeview.")

    # Bind the double-click event to the delete_selected_loan function
    result_tree.bind("<Double-1>", delete_selected_loan)

    def search_current_loans(query=""):
        try:
            # Load the workbook
            wb = load_workbook('database.xlsx')
            loans_sheet = wb['Current_Loans']
            student_sheet = wb['Student']
            instruments_sheet = wb['Instruments']

            # Clear the result tree
            result_tree.delete(*result_tree.get_children())

            # Create a dictionary to map student IDs to their names
            student_dict = {}
            for row in student_sheet.iter_rows(min_row=2, values_only=True):
                student_name = row[0]  # Assuming the student name is in the first column
                student_id = row[10]  # Assuming the student ID is in the 11th column
                student_dict[student_id] = student_name

            # Create a dictionary to map instrument IDs to their names
            instrument_dict = {}
            for row in instruments_sheet.iter_rows(min_row=2, values_only=True):
                instrument_id = row[1]  # Assuming the instrument ID is in the second column
                instrument_name = row[0]  # Retrieve the instrument name from column A
                instrument_dict[instrument_id] = instrument_name

            # Search for matching loans
            for row in loans_sheet.iter_rows(min_row=2, values_only=True):
                pupil_id, instrument_id, date_of_loan, _, _, date_of_return = row[:6]

                # Get the student name from the dictionary
                student_name = student_dict.get(pupil_id, f"Unknown - ID: {pupil_id}")

                # Get the instrument name from the dictionary
                instrument_name = instrument_dict.get(instrument_id, f"Unknown - ID: {instrument_id}")

                # Convert date_of_loan to the desired string format
                date_of_loan_str = date_of_loan[:2] + "/" + date_of_loan[2:4] + "/" + date_of_loan[4:]

                # Convert date_of_return to the desired string format if it exists
                if date_of_return:
                    date_of_return_str = date_of_return[:2] + "/" + date_of_return[2:4] + "/" + date_of_return[4:]
                else:
                    date_of_return_str = "NONE"

                # Check if the loan matches the search query
                if (
                    query.lower() in student_name.lower()
                    or query.lower() in instrument_name.lower()
                    or query.lower() in date_of_loan_str.lower()
                    or query.lower() in date_of_return_str.lower()
                ):
                    # Add the loan details to the result tree
                    result_tree.insert("", END, values=(student_name, instrument_name, date_of_loan_str, date_of_return_str))

        except PermissionError:
            messagebox.showerror("Error", "Cannot access the database. Please close the database file (database.xlsx) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    # Perform the initial search
    search_current_loans()

    def raise_back_button(event):
        back_button.lift()

    root.bind("<Configure>", raise_back_button)

def clear_window(root):
    for widget in root.winfo_children():
        if isinstance(widget, tk.Toplevel):
            widget.destroy()
        else:
            try:
                widget.destroy()
            except tk.TclError:
                pass

    # Unbind all events from the root window
    root.unbind_all("<Key>")
    root.unbind_all("<Button>")
    root.unbind_all("<Motion>")
    root.unbind_all("<<Event>>")

    # Remove all bindings for the <<ClearWindow>> event
    root.event_delete("<<ClearWindow>>")

    # Generate the <<ClearWindow>> event to trigger any remaining event handlers
    root.event_generate("<<ClearWindow>>")

    # Explicitly unbind the <<Configure>> event
    root.unbind("<Configure>")

def create_back_button(root, back_command):
    back_button = Button(root, text="Back", command=back_command, font=("Arial", 14))
    back_button.pack(padx=10, pady=10, anchor="nw")  # anchor="nw" places the widget in the top-left corner
    return back_button

def go_back(root):
    # If there are more than one page in the stack, go back to the previous page
    if len(page_stack) > 1:
        page_stack.pop()  # Remove the current page from the stack
        previous_page = page_stack.pop()  # Get the previous page
        previous_page(root)  # Show the previous page
    else:
        # If there is only one page in the stack (homepage), do nothing
        pass

def form_sorting_key(form):
    match = re.match(r"([LU]?)(\d+)([a-zA-Z]?)", form)
    if match:
        letter1, year, letter2 = match.groups()
        year = int(year)
        if not letter1:
            letter1 = "L"  # Treat as lower if the first letter is missing
        if not letter2:
            letter2 = "a"
        return (letter1 == "L", year, letter1, letter2.lower())
    else:
        return (True, float('inf'), "", "")  # Put non-matching forms at the end

if __name__ == "__main__":
    root = Tk()
    homepage(root)
    root.mainloop()